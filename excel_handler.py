# -*- coding: utf-8 -*-


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sys
import os
import threading
import multiprocessing

from data_handler import DataHandler

reload(sys)
sys.setdefaultencoding('utf-8')

db_conf = {
	"host": "xxxxx",
	"port": "xxx",
	"user": "xxxxx",
	"passwd": "xxxxx",
	"db": "bitnami_redmine",
	"charset": "utf8mb4"
}

WORKBOOK_NAME = "MOA迭代统计数据.xlsx"

'''================================================================'''
### 只用于多进程统计
# import copy_reg
# import types
 
# def _pickle_method(m):
#     if m.im_self is None:
#         return getattr, (m.im_class, m.im_func.func_name)
#     else:
#         return getattr, (m.im_self, m.im_func.func_name)
 
# copy_reg.pickle(types.MethodType, _pickle_method)

def _proxy(instance, event, *args):
	return getattr(instance, event)(*args)

'''================================================================'''

class ExcelHandler(object):
	"""将统计数据写入excel文件"""
	def __init__(self):
		self.data_handler = DataHandler(**db_conf)

	def create_static_workbook_sheet(self):
		''' @summary: 创建excel文件及相关信息 '''

		if os.path.isfile(WORKBOOK_NAME):
			os.remove(WORKBOOK_NAME)

		self.wb = Workbook()
		self.person_version_bug_sheet = self.wb.active
		self.person_version_bug_sheet.title = unicode("个人迭代bug数")
		self.module_version_bug_sheet = self.wb.create_sheet()
		self.module_version_bug_sheet.title = unicode("模块迭代bug数")
		self.person_version_work_weight_sheet = self.wb.create_sheet()
		self.person_version_work_weight_sheet.title = unicode("个人迭代工作粒度总和")
		self.person_not_finish_issue_sheet = self.wb.create_sheet()
		self.person_not_finish_issue_sheet.title = unicode("个人未完成任务数")
		self.person_online_issue_sheet = self.wb.create_sheet()
		self.person_online_issue_sheet.title = unicode("个人迭代网上问题数")
		self.web_person_version_work_weight_sheet = self.wb.create_sheet()
		self.web_person_version_work_weight_sheet.title = unicode("web个人迭代工作粒度总和")
		self.__save_workbook()

	def write_static_data(self, start_time, end_time):
		''' @summary 单线程实现数据统计 
		    @note: 这些方法都接收一个项目id的参数
		           1：MOA迭代项目
		           3：web组迭代项目'''

		self.__write_person_version_bug(1)
		self.__write_module_version_bug(1)
		self.__write_person_online_issue(start_time, end_time, 1)
		self.__write_person_not_finish_issue(1)
		self.__write_person_version_work_weight(1)
		self.__write_web_person_version_work_weight(3)

	def write_static_data_multiThread(self, start_time, end_time):
		''' @summary: 多线程实现数据统计
		    @note: 由于Python的GIL机制，多线程不一定效率更高 
		    该方法没有测试，应该是没有问题的 '''

		events = [
			self.__write_person_version_bug,
			self.__write_module_version_bug,
			self.__write_person_online_issue,
			self.__write_person_not_finish_issue,
			self.__write_person_version_work_weight
		]

		for event in events:
			t = None
			if event == self.__write_person_online_issue:
				t = threading.Thread(target = event, args = (start_time, end_time))
			else:
				t = threading.Thread(target = event)

			t.start()
			t.join()

	def write_static_data_multiProcess(self, start_time, end_time):
		''' @summary: 多进程实现数据统计 
			@note: 注意pickle机制
			该方法没有测试，应该是没有问题的'''

		events = [
			"_ExcelHandler__write_person_version_bug",
			"_ExcelHandler__write_module_version_bug",
			"_ExcelHandler__write_person_online_issue",
			"_ExcelHandler__write_person_not_finish_issue",
			"_ExcelHandler__write_person_version_work_weight"
		] #双下划线的属性在类中实际的属性名是需要加个类名前缀的

		#注意：进程数不要太多
		# cpus = multiprocessing.cpu_count() #获取CPU个数
		pool = multiprocessing.Pool(processes = len(events))
		results = []

		for event in events:
			# _pickle_method(event)

			result = None
			if event == "_ExcelHandler__write_person_online_issue":
				event = _proxy(self, event, start_time, end_time)
				result = pool.apply_async(event, args = (start_time, end_time))
			else:
				event = _proxy(self, event)
				result = pool.apply_async(event)
			results.append(result)

		pool.close()
		pool.join()

		# for result in results:
		# 	print("result: ", result.get())

  		print("done")

	def __save_workbook(self):
		self.wb.save(filename = WORKBOOK_NAME)	

	'''================================================================'''
	def __write_proj_version_row(self, sheet, proj_id):
		""" 写入excel文档的版本行：第1行，第2列开始 """
		proj_versions_info_sorted = self.__get_proj_versions_list_sorted(proj_id)
		for index in xrange(0, len(proj_versions_info_sorted)):
			sheet.cell(column=index+2, row=1, value=proj_versions_info_sorted[index]["name"])
		self.__save_workbook()

	def __write_username_col(self, sheet):
		""" 写入excel文档的用户列：第1列，第2行开始 """
		users_info_sorted = self.__get_users_list_sorted()
		for index in xrange(0, len(users_info_sorted)):
			sheet.cell(column=1, row=index+2, value=users_info_sorted[index]["name"])
		self.__save_workbook()

	def __write_module_name_col(self, sheet):
		""" 写入excel文档的模块列：第1列，第2行开始 """
		modules_name_sorted = self.__get_modules_name_list_sorted()
		for index in xrange(0, len(modules_name_sorted)):
			sheet.cell(column=1, row=index+2, value=modules_name_sorted[index])
		self.__save_workbook()

	'''================================================================'''
	def __get_proj_versions_list_sorted(self, proj_id):
		""" 返回：排好序的项目版本列表 """
		proj_versions_info = self.data_handler.get_proj_versions_list(proj_id)
		proj_versions_info_sorted = self.__list_contain_dict_sort(proj_versions_info, "id")
		print("proj_versions_info: {}".format(proj_versions_info_sorted))
		return proj_versions_info_sorted

	def __get_users_list_sorted(self):
		""" 返回：排好序的用户列表 """
		users_info = self.data_handler.get_users_list()
		users_info_sorted = self.__list_contain_dict_sort(users_info, "id")
		print("users_info: {}".format(users_info_sorted))
		return users_info_sorted

	def __list_contain_dict_sort(self, source_data, sort_key):
		""" list内嵌套dict排序: 根据list中dict元素的sort_key升序"""
		return sorted(source_data, key = lambda item: item[sort_key])

	'''================================================================'''
	def __get_modules_name_list_sorted(self):
		""" 返回：排好序的模块name列表 """
		modules_name = self.data_handler.get_proj_modules_name_list()
		modules_name_sorted = sorted(modules_name)
		for name in modules_name_sorted:
			print("module_name: {}".format(name))
		return modules_name_sorted

	'''======================================================================='''
	def __write_person_version_bug(self, proj_id):
		""" 写入个人迭代bug数excel文档 """
		print("write_person_bug")
		self.__write_proj_version_row(self.person_version_bug_sheet, proj_id)
		self.__write_username_col(self.person_version_bug_sheet)

		users_info_sorted = self.__get_users_list_sorted()
		proj_versions_info_sorted = self.__get_proj_versions_list_sorted(proj_id)
		for u_index in xrange(0, len(users_info_sorted)):
			for v_index in xrange(0, len(proj_versions_info_sorted)):
				person_version_bug_count = self.data_handler.get_person_version_bug_count(users_info_sorted[u_index]["name"], proj_versions_info_sorted[v_index]["name"], proj_id)
				if person_version_bug_count == 0:
					continue
				self.person_version_bug_sheet.cell(column=v_index+2, row=u_index+2, value=person_version_bug_count)
		self.__save_workbook()

	'''======================================================================='''
	def __write_module_version_bug(self, proj_id):
		""" 写入模块迭代bug数excel文档 """
		self.__write_proj_version_row(self.module_version_bug_sheet, proj_id)
		self.__write_module_name_col(self.module_version_bug_sheet)

		modules_name_sorted = self.__get_modules_name_list_sorted()
		proj_versions_info_sorted = self.__get_proj_versions_list_sorted(proj_id)
		for m_index in xrange(0, len(modules_name_sorted)):
			for v_index in xrange(0, len(proj_versions_info_sorted)):
				module_version_bug_count = self.data_handler.get_module_version_bug_count(modules_name_sorted[m_index], proj_versions_info_sorted[v_index]["name"], proj_id)
				if module_version_bug_count == 0:
					continue
				self.module_version_bug_sheet.cell(column=v_index+2, row=m_index+2, value=module_version_bug_count)
		self.__save_workbook()

	'''======================================================================='''
	def __write_person_online_issue(self, start_time, end_time, proj_id):
		""" 写入个人网上问题数excel文档 """
		self.__write_username_col(self.person_online_issue_sheet)
		self.person_online_issue_sheet.cell(column=2, row=1, value="{}至{}".format(start_time, end_time))

		users_info_sorted = self.__get_users_list_sorted()
		for u_index in xrange(0, len(users_info_sorted)):
			person_online_count = self.data_handler.get_person_online_issue_count(users_info_sorted[u_index]["name"], start_time, end_time, proj_id)
			if person_online_count == 0:
				continue
			self.person_online_issue_sheet.cell(column=2, row=u_index+2, value=person_online_count)
		self.__save_workbook()

	'''======================================================================='''
	def __write_person_not_finish_issue(self, proj_id):
		""" 写入个人遗留问题数excel文档 """
		self.__write_username_col(self.person_not_finish_issue_sheet)
		self.person_not_finish_issue_sheet.cell(column=2, row=1, value="遗留问题")

		users_info_sorted = self.__get_users_list_sorted()
		for u_index in xrange(0, len(users_info_sorted)):
			person_not_finish_count = self.data_handler.get_person_not_finish_issue_count(users_info_sorted[u_index]["name"], proj_id)
			if person_not_finish_count == 0:
				continue
			self.person_not_finish_issue_sheet.cell(column=2, row=u_index+2, value=person_not_finish_count)
		self.__save_workbook()

	'''======================================================================='''
	def __write_person_version_work_weight(self, proj_id):
		""" 写入个人迭代工作粒度excel文档 """
		self.__write_proj_version_row(self.person_version_work_weight_sheet, proj_id)
		self.__write_username_col(self.person_version_work_weight_sheet)

		users_info_sorted = self.__get_users_list_sorted()
		proj_versions_info_sorted = self.__get_proj_versions_list_sorted(proj_id)
		for u_index in xrange(0, len(users_info_sorted)):
			for v_index in xrange(0, len(proj_versions_info_sorted)):
				person_version_work_weight = self.data_handler.get_person_version_work_weight(users_info_sorted[u_index]["name"], proj_versions_info_sorted[v_index]["name"], proj_id)
				if person_version_work_weight == 0:
					continue
				self.person_version_work_weight_sheet.cell(column=v_index+2, row=u_index+2, value=person_version_work_weight)
		self.__save_workbook()

	'''======================================================================='''
	def __write_web_person_version_work_weight(self, proj_id):
		""" 写入web个人迭代工作粒度excel文档 """
		self.__write_proj_version_row(self.web_person_version_work_weight_sheet, proj_id)
		self.__write_username_col(self.web_person_version_work_weight_sheet)

		users_info_sorted = self.__get_users_list_sorted()
		proj_versions_info_sorted = self.__get_proj_versions_list_sorted(proj_id)
		for u_index in xrange(0, len(users_info_sorted)):
			for v_index in xrange(0, len(proj_versions_info_sorted)):
				person_version_work_weight = self.data_handler.get_person_version_work_weight(users_info_sorted[u_index]["name"], proj_versions_info_sorted[v_index]["name"], proj_id)
				if person_version_work_weight == 0:
					continue
				self.web_person_version_work_weight_sheet.cell(column=v_index+2, row=u_index+2, value=person_version_work_weight)
		self.__save_workbook()

	# def __set_head_cell_style(self, col, row, style):
	# 	col_name = get_column_letter(col)
	# 	cell_pos = "{}{}".format(col_name, row)
	# 	cell = self.wb[cell_pos]
	# 	cell.font = style



		
